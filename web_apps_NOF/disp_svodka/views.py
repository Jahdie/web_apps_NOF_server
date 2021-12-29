from django.shortcuts import render
from django.template.loader import render_to_string
from disp_svodka.models import *
from django.http import JsonResponse, HttpResponse
from openpyxl.styles.borders import Border, Side
import openpyxl
import pypyodbc
from services import connecting_to_db, get_values_from_db, merging_thead_and_tbody


def index(request):
    context = {}
    return render(request, 'base.html')


def disp_svodka_controller(request):
    context = {'tabs': TabNames.objects.all()}

    return render(request, 'disp_svodka/base.html', context)


def get_table_of_dispatchers_summary_by_tab_name(request, date_selected, tab_name):
    if request.is_ajax():

        th = TableHeads.get_theads(date_selected=date_selected, tab_name=tab_name)
        ts = TableSiders.get_tsiders(date_selected=date_selected, tab_name=tab_name)
        table = merging_thead_and_tbody.merging_thead_and_tbody(th=th, ts=ts)
        wb = openpyxl.load_workbook('C:\\Users\\Stas\\OneDrive\\Рабочий стол\\test.xlsx')
        ws = wb.active

        border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))
        #
        for item in table.values():
            for a in item.values():
                for el in a:
                    column_start = el['column_start']
                    column_end = el['column_end']
                    row_start = el['row_start']
                    row_end = el['row_end']
                    value = el['field_value']

                    # print(column_start, column_end, row_start, row_end, el, value)

                    ws.cell(row=row_start, column=column_start).value = value
                    ws.cell(column=column_start, row=row_start).border = border
                    ws.merge_cells(start_column=column_start, end_column=column_end, start_row=row_start,
                                   end_row=row_end)

        wb.save('C:\\Users\\Stas\\OneDrive\\Рабочий стол\\test.xlsx')

        context = {'table': table}
        return render(request, 'disp_svodka/test.html', context)
